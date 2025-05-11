#  ----------------------------------------------更新日志---------------------------------------------------------------------
##  
##  2025/03/12  修复bug，perform_comparison中file1_name处理策略错误，使用resplit()后修复
##  2025/03/12  功能改进，增加index输入框识别字母，可以将excel的表头字母序数直接输入
##  2025/03/12  功能新增，增加mapping表头的功能选项
##  2025/03/23  已完成所有按钮添加，按钮功能已可以正常运行，下一步需要添加具体的按title的对比逻辑，需要调整的function = compare_excel_sheets_by_index
##  2025/03/24  已完成功能新增，增加mapping表头的功能
##  2025/03/24  代码优化，优化线程管理，避免在线程中对UI进行操作·····
##  2025/03/24  代码优化，完成优化线程管理
##  2025/03/24  功能增加，应用关闭是保存当前所有部件状态，并在下次打开时保持之前状态·····
##  2025/03/27  功能增加，完成应用关闭时保存当前所有部件状态
##  2025/03/27  新增一个一键清空所有部件内容的功能
##  2025/03/30  效率提升，对比时间减少70%
##  2025/04/08  修复index计算错误的问题
##  2025/03/30  index列各个列未合并时，如何处理?
##  2025/04/12  增加一个index列，使用两列合并值作为index.....
##  2025/04/12  从190行开始分析，解决self.check_index_repeat的问题，多列加在一起校验repeat，且需要考虑某个index为空值时的情况.....
##  2025/04/13  增加多列组合作为索引功能
##  2025/04/13  增加stop按钮，用于终止线程（当前无法立刻终止进程，会直接崩溃）.....已解决，待测试
##  2025/04/13  增加显示当前运行状态，或者细化进度条，步进改为1%.....
##  2025/04/13  764行，计算index值有问题，待解决.....
##  2025/04/15  764行，计算index值有问题，已解决
##  2025/04/19  因单元格格式问题，存在很多隐藏回车符，导致对比结果为not_agreed，以解决此问题
##  2025/04/19  增加文本框，用于展示当前正在执行的任务.....
##  2025/04/20  增加文本框，已完成
##  2025/04/22  优化单元格为空时的逻辑，提高效率.....
##  2025/04/22  完成优化单元格为空时的逻辑
##  2025/05/08  封装cell_consistency_check方法，简化代码
##  2025/05/08  解决浮点数对比不同问题，在字符串对比不同后增加浮点数对比，此问题原因是当文件有数据验证且数据验证有小数点时，会导致单元格内容转字符串时强制带小数，导致对比不同
##	
##  ----------------------------------------------未完待续---------------------------------------------------------------------

##
##
##

import openpyxl
import sys
import os
import inspect
import copy
import json
import ctypes
import time
import threading
from openpyxl.styles import PatternFill, Alignment
import openpyxl.styles
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QFileDialog, QDialog, QProgressBar, QPlainTextEdit
import openpyxl.workbook
from PyQt5.QtCore import Qt, QPropertyAnimation, QEasingCurve, QEvent, QThread, QThreadPool, QRunnable, pyqtSignal, QObject, Qt
from PyQt5.QtGui import QFont

if __name__ == "__main__":
    # 这里是直接运行模块时要执行的代码逻辑
    # package_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # sys.path.append(package_root)
    # from self_package import self_tool
    import self_tool
else:
    # 这里是模块被导入时要执行的代码逻辑，比如执行相关的函数、初始化操作等
    from ..self_package import self_tool

output_path_name = ".\outputfile"

class Person_ComparisonApp:
    # 类的构造函数，用于初始化对象的属性
    def __init__(self, name):
        self.output_path_name = name
        self.Progress_percent = 0
        self.Agreed_color       = "9AFF9A" #PaleGreen1
        self.Not_Agreed_color   = "E44C63" #Red
        self.No_match_color     = "00FFFF" #Cyan
        self.None_color         = "FFFFFF" #White
        self.is_running         =  True

    def check_index_repeat(self, comparisontask, sheet, index_column, sheet_name, file_path):
        # 检查重点对比的sheet中的索引列是否有重复元素。
        # :param comparisontask : 线程运行状态标志，用于终止编程；
        # :param sheet          : 当前工作表对象；
        # :param index_column   : 索引列的列号list；
        # :param sheet_name     : 工作表名称；
        # :param file_path      : 文件路径；
        # flag = False
        error_message = f"文档 : {file_path}\n"
        index_values = index_column.copy()
        while 0 in index_values:
            index_values.remove(0)
        
        blank_row_flag = 0   
        for i in range(1, sheet.max_row):
            if self.check_thread_running(comparisontask):
                return 0

            merged_text_major = self.get_merged_text(sheet, i, index_values)
            print(f"sheet{sheet_name} 的{i}行")
            comparisontask.signals.progress_current_task.emit(f"正在根据索引值对比sheet【{sheet_name}】中，检查index列的第{i}行与其后的行是否重复")
            if not merged_text_major:
                # 跳过合并文本为空的行
                blank_row_flag += 1
                if blank_row_flag >= 30:    #如果连续30行的index列合并后的值为None，那么认为下面全部为空，停止检查
                    print_info = f"连续30行全部值为空，结束sheet【{sheet_name}】的检查index列名称重复================"
                    print(print_info)
                    comparisontask.signals.progress_current_task.emit(print_info)
                    break
                else:
                    pass
            else:
                blank_row_flag = 0    #merged_text_major不为空时，flag置为0，需要全部连续才能使flag持续+1，否则清零
                for j in range(i + 1, sheet.max_row+1):
                    if self.check_thread_running(comparisontask):
                        return 0
                    merged_text_minor = self.get_merged_text(sheet, j, index_values)

                    if merged_text_major == merged_text_minor:
                        error_message += f"sheet：[{sheet_name}]第[{index_column}]列的第{i}行与第{j}行名称相同：【{merged_text_major}】\n"
                        print(error_message)
                        comparisontask.signals.progress_current_task.emit(f"正在根据索引值对比sheet【{sheet_name}】中，第[{index_column}]列合并后，第{i}行与第{j}行名称相同【{merged_text_major}】================")
                        ctypes.windll.user32.MessageBoxW(None, f"文档 : {file_path}\n sheet：[{sheet_name}]\n第[{index_column}]列合并后存在相同元素\n第{i}行与第{j}行名称相同：【{merged_text_major}】",
                                                        "错误信息", 0x00000010)
                        return 0
                    

        # if flag:
        #     print(error_message)
        #     ctypes.windll.user32.MessageBoxW(None, f"文档 : {file_path}\n sheet：[{sheet_name}]\n第[{index_column}]列合并后存在相同元素",
        #                                     "错误信息", 0x00000010)
        #     return 0
        # print(f"文档 : {file_path}的第{index_column}无重复项\n")
        return 1
    # 拆分合并单元格，并复制格式信息。

    def get_merged_text(self, sheet, row, index_values):
            """
            优化点：
            1. 将字符串替换操作提前到获取单元格值时进行，避免在后续比较中重复操作。
            sheet           : 待合并列值的sheet
            row             : 待合并列值的行
            index_values    : 字典，存储待合并的列数

            """
            merged_text = ""
            for col in index_values:
                value = sheet.cell(row, col).value
                if value is not None:
                    # 提前处理字符串，避免后续重复替换
                    value = str(value).replace('_x000D_', '').replace('\r', '').replace('\n', '').replace(' ', '')
                    merged_text += value
            return merged_text

    def copy_cell_format(self, sheet, row1, col1, row_min, col_min):
        """
        优化点：
        1. 将单元格格式复制操作封装成独立函数，提高代码复用性。
        sheet     ：待填充单元格的sheet
        row1      ：当前被拆分单元格的行数，待设置格式的单元格
        col1      ：当前被拆分单元格的列数，待设置格式的单元格
        row_min   ：左上角单元格的行数，被赋值格式的左上角单元格
        col_min   ：左上角单元格的列数，被赋值格式的左上角单元格
        """
        if sheet.cell(row=row1, column=col1).fill:
            start_color = sheet.cell(row=row_min, column=col_min).fill.start_color
            end_color = sheet.cell(row=row_min, column=col_min).fill.end_color
            fill_type = sheet.cell(row=row_min, column=col_min).fill.fill_type
            sheet.cell(row=row1, column=col1).fill = PatternFill(start_color=start_color, end_color=end_color, fill_type=fill_type)
        if sheet.cell(row=row1, column=col1).border:
            left_border = sheet.cell(row=row_min, column=col_min).border.left
            right_border = sheet.cell(row=row_min, column=col_min).border.left
            top_border = sheet.cell(row=row_min, column=col_min).border.top
            bottom_border = sheet.cell(row=row_min, column=col_min).border.bottom
            sheet.cell(row=row1, column=col1).border = openpyxl.styles.Border(left=left_border, right=right_border, top=top_border, bottom=bottom_border)
        if sheet.cell(row=row1, column=col1).font:
            font_name = sheet.cell(row=row_min, column=col_min).font.name
            font_size = sheet.cell(row=row_min, column=col_min).font.size
            font_color = sheet.cell(row=row_min, column=col_min).font.color
            font_bold = sheet.cell(row=row_min, column=col_min).font.bold
            font_italic = sheet.cell(row=row_min, column=col_min).font.italic
            font_underline = sheet.cell(row=row_min, column=col_min).font.underline
            sheet.cell(row=row1, column=col1).font = openpyxl.styles.Font(name=font_name, size=font_size, color=font_color, bold=font_bold, italic=font_italic, underline=font_underline)
        font_horizontal = sheet.cell(row=row_min, column=col_min).alignment.horizontal
        font_vertical = sheet.cell(row=row_min, column=col_min).alignment.vertical
        font_wrap_text = sheet.cell(row=row_min, column=col_min).alignment.wrap_text
        sheet.cell(row=row1, column=col1).alignment = openpyxl.styles.Alignment(horizontal=font_horizontal, vertical=font_vertical, wrap_text=font_wrap_text)
        sheet.cell(row=row1, column=col1).number_format = sheet.cell(row=row_min, column=col_min).number_format

    def split_merged_cells(self, comparisontask, wb1, name, compare_col_index):
        sheet = wb1[name]
        merged_cell_ranges = list(sheet.merged_cells.ranges)
        for merged_cell in merged_cell_ranges:
            row_min = merged_cell.min_row
            row_max = merged_cell.max_row
            col_min = merged_cell.min_col
            col_max = merged_cell.max_col
            sheet.unmerge_cells(start_row=row_min, start_column=col_min, end_row=row_max, end_column=col_max)
            for row1 in range(row_min, row_max+1):
                if self.check_thread_running(comparisontask):
                    return 0
                for col1 in range(col_min, col_max+1):
                    if self.check_thread_running(comparisontask):
                        return 0
                    if not(row1 == row_min and col1 == col_min):
                        index_row = row1 - row_min
                        index_col = col1 - col_min
                        sheet.cell(row=row1, column=col1).value = sheet.cell(row=row_min, column=col_min).value
                        # 判断是否是索引列
                        if col1 in compare_col_index: # 单元格在索引列，赋值必须唯一{左上角值+行值+列值}
                            sheet.cell(row=row1, column=col1).value = f"{sheet.cell(row=row1, column=col1).value}{index_row}{index_col}"
                        else: # 单元格不在索引列，赋值统一为左上角值
                            sheet.cell(row=row1, column=col1).value = f"{sheet.cell(row=row1, column=col1).value}"
                        # 复制单元格格式
                        self.copy_cell_format(sheet, row1, col1, row_min, col_min)

    def open_file(self, file1_path):
        # 加载一个 Excel 文件
        try:
            wb = openpyxl.load_workbook(file1_path)
        except FileNotFoundError:
            error = f"文件 {file1_path} 不存在。"
            print(error)
            ctypes.windll.user32.MessageBoxW(None, error, "错误信息", 0x00000010)
            return 0
        except openpyxl.utils.exceptions.InvalidFileException:
            error = f"文件 {file1_path} 不是有效的 Excel 文件, 请重新输入"
            print(error)
            ctypes.windll.user32.MessageBoxW(None, error, "错误信息", 0x00000010)
            return 0
        except Exception as e:
            error = f"发生了未知错误：{e}"
            print(error)
            ctypes.windll.user32.MessageBoxW(None, error, "错误信息", 0x00000010)
            return 0
        return wb
    
    def cell_consistency_check(self, sheet1_cell, sheet2_cell):
        sheet1_cell.alignment = Alignment(wrap_text=True)    #把第一个文件的单元格设为自动换行
        sheet2_cell.alignment = Alignment(wrap_text=True)   #把第二个文件的单元格设为自动换行
        value1 = str(sheet1_cell.value).replace('_x000D_', '').replace('\r', '').replace('\n', '').replace(' ', '')  #sheet1中对应单元格的值
        value2 = str(sheet2_cell.value).replace('_x000D_', '').replace('\r', '').replace('\n', '').replace(' ', '')  #sheet2中对应单元格的值
        if str(value1) == "None":
            value1 = ""
        if str(value2) == "None":
            value2 = ""

        if value1 == value2:  # 单元格值对比相同
            return True
        else:
            try:
                # 尝试将单元格的值转换为浮点数
                value3 = float(sheet1_cell.value)
                value4 = float(sheet2_cell.value)
                # 比较两个浮点数
                if value3 == value4:
                    return True
                else:
                    return False
            except (ValueError, TypeError):
                # 如果无法转换为浮点数
                return False
    
    def compare_excel_sheets_by_index(self, comparisontask, wb1, wb2, output_path, compare_sheet_name, number):
        index_column_number = 0
        # 检查两个文件中分别是否包含compare_sheet_name中的每个键值，以及索引值是否超出该sheet的max_column
        # print(f"compare_excel_sheets_by_index : compare_sheet_name = {compare_sheet_name}")
        for name in compare_sheet_name:
            comparisontask.signals.progress_current_task.emit(f"检查按索引对比的sheet名称是否存在于文件中，index值是否超范围···")
            if self.check_thread_running(comparisontask):
                return 0
            index_column_number = len(compare_sheet_name[name]) - 1 #获取索引列个数 index_column_number = 索引列个数，---最后一个元素是mapping flag：0=no mapping 1=mapping
            if name in wb1.sheetnames and name in wb2.sheetnames:
                for x in range(0, index_column_number):
                    if compare_sheet_name[name][x] > wb1[name].max_column or compare_sheet_name[name][x] > wb2[name].max_column: # or compare_sheet_name[name][x] == 0:
                        if wb1[name].max_column > wb2[name].max_column:
                            error_message1 = f"sheet name输入错误：sheet[{name}]中的第{compare_sheet_name[name][x]}列超出wb2的最大列范围{wb2[name].max_column}\n"
                        else:
                            error_message1 = f"sheet name输入错误：sheet[{name}]中的第{compare_sheet_name[name][x]}列超出wb1的最大列范围{wb1[name].max_column}\n"
                        print(error_message1)
                        ctypes.windll.user32.MessageBoxW(None, error_message1, "错误信息", 0x00000010)
                        return 0
            else:
                error_message2 = f"sheet name输入错误：sheet[{name}]在其中一个文档中不存在\n"
                print(error_message2)
                ctypes.windll.user32.MessageBoxW(None, error_message2, "错误信息", 0x00000010)
                return 0
        total_sheets = len(wb1.sheetnames)+len(wb2.sheetnames)
        for sheet_name in wb1.sheetnames:
            if self.check_thread_running(comparisontask):
                return 0
            self.Progress_percent += int(100/total_sheets)
            comparisontask.signals.progress_update.emit(self.Progress_percent)
            comparisontask.signals.progress_current_task.emit(f"对比sheet {sheet_name}")
            # ComparisonApp.compare_progress_bar.setValue(self.Progress_percent)
            if sheet_name in wb2.sheetnames:
                # 获取第一个文件中的工作表
                sheet1 = wb1[sheet_name]
                # 获取第二个文件中的工作表
                sheet2 = wb2[sheet_name]
                #index_col = index_column  # D 列是第 4 列
                max_row1 = sheet1.max_row
                max_row2 = sheet2.max_row
                max_col1 = sheet1.max_column
                max_col2 = sheet2.max_column
                # # break_col = [0]
                # sheet1.data_validations.dataValidation = []
                # print(f"file1 sheet【{sheet_name}】的数据验证已清除")
                # comparisontask.signals.progress_current_task.emit(f"file1 sheet【{sheet_name}】的数据验证已清除")
                # sheet2.data_validations.dataValidation = []
                # print(f"file2 sheet【{sheet_name}】的数据验证已清除")
                # comparisontask.signals.progress_current_task.emit(f"file2 sheet【{sheet_name}】的数据验证已清除")


                # compare_sheet_name不为空，且文件中存在这个sheet
                if not(compare_sheet_name == None) and sheet_name in compare_sheet_name:
                    print(f"正在根据索引值对比sheet【{sheet_name}】中")
                    name = f'''{sheet_name}'''
                    print(f"compare_excel_sheets_by_index : compare_sheet_name = {compare_sheet_name}")
                    print(f"compare_excel_sheets_by_index : sheet_name = {sheet_name}")
                    index_values = compare_sheet_name[sheet_name].copy()
                    mapping_flag = index_values.pop() #获取最后一个元素，mapping_flag：0=no mapping 1=mapping
                    while 0 in index_values:
                        index_values.remove(0)
                    print(f"compare_excel_sheets_by_index: index_values = {index_values}")
                    # 检查是否有重复单元格
                    message = output_path.split('\\')[-1].replace("-compare", "")
                    # print(f"output_path.split('\\')[-1].replace(, "")  = {message}")
                    if number == 0:
                        comparisontask.signals.progress_current_task.emit(f"正在根据索引值对比sheet【{sheet_name}】中，检查index列的每行之间是否存在重复值")
                        if self.check_index_repeat(comparisontask, sheet1, index_values, sheet_name, message) == 0:
                            return 0
                        if self.check_index_repeat(comparisontask, sheet2, index_values, sheet_name, message) == 0:
                            return 0
                    # 拆分合并单元格
                    comparisontask.signals.progress_current_task.emit(f"正在根据索引值对比sheet【{sheet_name}】中，拆分File1的合并单元格")
                    self.split_merged_cells(comparisontask, wb1, sheet_name, index_values)
                    comparisontask.signals.progress_current_task.emit(f"正在根据索引值对比sheet【{sheet_name}】中，拆分File2的合并单元格")
                    self.split_merged_cells(comparisontask, wb2, sheet_name, index_values)
                    print(f"compare_excel_sheets_by_index: self.split_merged_cells")
                    index_column_mapping = {}
                    title_row_mapping = {}
                    # mapping index列的所有行
                    blank_row_flag = 0
                    for row1 in range(2, max_row1 + 1): #遍历第一个文件每一行
                        comparisontask.signals.progress_current_task.emit(f"正在根据索引值对比的sheet【{sheet_name}】中，查找File1第{row1}行对应File2的行···")
                        if self.check_thread_running(comparisontask):
                            return 0
                        #如果未mapping上，将保持0
                        index_column_mapping[row1] = 0
                        # 获取第一份文件当前行的 索引列 的值
                        print(f"compare_excel_sheets_by_index: 遍历第一个文件第{row1}行")
                        merged_text_major = self.get_merged_text(sheet1, row1, index_values)
                        # value1 = sheet1.cell(row=row1, column=compare_sheet_name[name][0]).value
                        if merged_text_major != "" and str(merged_text_major) != "None":
                            blank_row_flag = 0
                            for row2 in range(2, max_row2 + 1): #遍历第二个文件的每一行
                                # comparisontask.signals.progress_current_task.emit(f"正在根据索引值对比的sheet【{sheet_name}】中，检查File1第{row1}行与File2的{row2}行是否匹配")
                                if self.check_thread_running(comparisontask):
                                    return 0
                                merged_text_minor = self.get_merged_text(sheet2, row2, index_values)
                                # value2 = sheet2.cell(row=row2, column=compare_sheet_name[name][0]).value
                                if merged_text_major == merged_text_minor:
                                    index_column_mapping[row1] = row2
                                else:
                                    continue
                            if index_column_mapping[row1] == 0:
                                for col in range(1, max_col1 + 1):
                                    cell = sheet1.cell(row=row1, column=col)
                                    cell.fill = PatternFill(start_color=self.No_match_color, end_color=self.No_match_color, fill_type="solid")
                        else:
                            for col in range(1, max_col1 + 1):
                                cell = sheet1.cell(row=row1, column=col)
                                cell.fill = PatternFill(start_color=self.No_match_color, end_color=self.No_match_color, fill_type="solid")
                            blank_row_flag += 1
                            if blank_row_flag >= 20:
                                print_info = f"连续20行全部值为空，结束查找sheet【{sheet_name}】的File1的行对应File2的行================"
                                print(print_info)
                                comparisontask.signals.progress_current_task.emit(print_info)
                                break
                            else:
                                pass

                    if mapping_flag != 0:
                    #---------------------需要mapping表头对比---------------------#
                        print(f"#---------------------需要mapping表头对比---------------------#{sheet_name}")
                        # mapping title行的所有列
                        for col1 in range(1, max_col1 + 1): #遍历第一个文件每一列
                            comparisontask.signals.progress_current_task.emit(f"正在根据索引值并mapping tile对比sheet【{sheet_name}】中，查找File1第{col1}列对应File2的列")
                            if self.check_thread_running(comparisontask):
                                return 0
                            #如果未mapping上，将保持0
                            title_row_mapping[col1] = 0
                            # 获取第一份文件当前行的 索引列 的值
                            value1 = str(sheet1.cell(row=1, column=col1).value).replace('_x000D_', '').replace('\r', '').replace('\n', '').replace(' ', '')
                            if value1 != "" and str(value1) != "None":
                                for col2 in range(1, max_col2 + 1): #遍历第二个文件的每一列
                                    # comparisontask.signals.progress_current_task.emit(f"正在根据索引值并mapping tile对比sheet【{sheet_name}】中，检查File1第{col1}列title与File2的{col2}列title是否匹配")
                                    if self.check_thread_running(comparisontask):
                                        return 0
                                    value2 = str(sheet2.cell(row=1, column=col2).value).replace('_x000D_', '').replace('\r', '').replace('\n', '').replace(' ', '')
                                    if value1 == value2:
                                        title_row_mapping[col1] = col2
                                        break

                                if title_row_mapping[col1] == 0:
                                    for row in range(1, max_row1 + 1):
                                        cell = sheet1.cell(row=row, column=col1)
                                        cell.fill = PatternFill(start_color=self.No_match_color, end_color=self.No_match_color, fill_type="solid")
                            else:
                                # for row in range(1, max_row1 + 1):
                                #     cell = sheet1.cell(row=row, column=col1)
                                #     cell.fill = PatternFill(start_color=self.No_match_color, end_color=self.No_match_color, fill_type="solid")
                                pass

                        #开始对比
                        blank_row_flag = 0   
                        for row1 in range(2, max_row1 + 1): #遍历第一个文件每一列
                            comparisontask.signals.progress_current_task.emit(f"正在根据索引值并mapping tile对比sheet【{sheet_name}】中，开始对比第{row1}行")
                            if self.check_thread_running(comparisontask):
                                return 0
                            # 获取第一份文件当前行的 索引列 的值
                            if index_column_mapping[row1] != 0:
                                blank_row_flag = 0   
                                # print(f"当前行数为：{inspect.currentframe().f_lineno}index_column_mapping[row1] = {index_column_mapping[row1]}，row1 = {row1}")
                                if self.check_thread_running(comparisontask):
                                    return 0
                                for col1 in range(1, max_col1 + 1): #遍历第一个文件第row1行的每一列
                                    # comparisontask.signals.progress_current_task.emit(f"正在根据索引值并mapping tile对比sheet【{sheet_name}】中，开始对比第{row1}行的第{col1}列")
                                    if title_row_mapping[col1] != 0:
                                        # print(f"当前行数为：{inspect.currentframe().f_lineno}title_row_mapping[col1] = {title_row_mapping[col1]}")
                                        sheet1_cell = sheet1.cell(row=row1, column=col1)    #获取第一个文件的单元格
                                        sheet2_cell = sheet2.cell(row=index_column_mapping[row1], column=title_row_mapping[col1])   #获取第二个文件的单元格
                                        if self.cell_consistency_check(sheet1_cell, sheet2_cell):
                                            sheet1_cell.fill = PatternFill(start_color=self.Agreed_color, end_color=self.Agreed_color, fill_type="solid")
                                        else:
                                            sheet1_cell.fill = PatternFill(start_color=self.Not_Agreed_color, end_color=self.Not_Agreed_color, fill_type="solid")
                                            print(f"当前行数为：{inspect.currentframe().f_lineno}，value1 = {sheet1_cell.value}")
                                            print(f"当前行数为：{inspect.currentframe().f_lineno}，value2 = {sheet2_cell.value}")
                                    else:
                                        continue
                            else:
                                blank_row_flag += 1
                                if blank_row_flag >= 20:
                                    print_info = f"连续20行全部值为空，结束sheet【{sheet_name}】的对比================"
                                    print(print_info)
                                    comparisontask.signals.progress_current_task.emit(print_info)
                                    break
                                else:
                                    pass
        
                    else:
                        #---------------------不需要mapping表头---------------------#    
                        print(f"#---------------------不需要mapping表头对比---------------------#{sheet_name}")
                        blank_col_flag = 0
                        blank_row_flag = 0
                        for row1 in range(2, max_row1 + 1): #遍历第一个文件每一行
                            comparisontask.signals.progress_current_task.emit(f"正在根据索引值对比sheet【{sheet_name}】中，对比第{row1}行")
                            if self.check_thread_running(comparisontask):
                                return 0
                            # 获取第一份文件当前行的 索引列 的值
                            if index_column_mapping[row1] != 0:
                                blank_row_flag = 0

                                for col1 in range(1, max_col1 + 1):
                                    if self.check_thread_running(comparisontask):
                                        return 0
                                    sheet1_cell = sheet1.cell(row=row1, column=col1)    #获取第一个文件的单元格
                                    sheet2_cell = sheet2.cell(row=index_column_mapping[row1], column=col1)   #获取第二个文件的单元格
                                    if self.cell_consistency_check(sheet1_cell, sheet2_cell):
                                        sheet1_cell.fill = PatternFill(start_color=self.Agreed_color, end_color=self.Agreed_color, fill_type="solid")
                                    else:
                                        sheet1_cell.fill = PatternFill(start_color=self.Not_Agreed_color, end_color=self.Not_Agreed_color, fill_type="solid")
                                        print(f"当前行数为：{inspect.currentframe().f_lineno}，value1 = {sheet1_cell.value}")
                                        print(f"当前行数为：{inspect.currentframe().f_lineno}，value2 = {sheet2_cell.value}")
                            else:
                                blank_row_flag += 1
                                if blank_row_flag >= 20:
                                    print_info = f"连续20行全部值为空，结束sheet【{sheet_name}】的对比================"
                                    print(print_info)
                                    comparisontask.signals.progress_current_task.emit(print_info)
                                    break
                                else:
                                    pass
                                                
                else:
                    blank_col_flag = 0
                    blank_row_flag = 0
                    print(f"直接对比: sheet_name  = {sheet_name}")
                    for row1 in range(1, max_row1 + 1):
                        blank_col_flag = 0
                        comparisontask.signals.progress_current_task.emit(f"直接对比sheet【{sheet_name}】，对比第{row1}行")
                        if self.check_thread_running(comparisontask):
                            return 0
                        for col in range(1, sheet1.max_column + 1):
                            if self.check_thread_running(comparisontask):
                                return 0
                            sheet1_cell = sheet1.cell(row=row1, column=col)
                            sheet2_cell = sheet2.cell(row=row1, column=col)
                            if self.cell_consistency_check(sheet1_cell, sheet2_cell):
                                sheet1_cell.fill = PatternFill(start_color=self.Agreed_color, end_color=self.Agreed_color, fill_type="solid")
                                value1 = sheet1_cell.value
                                value2 = sheet2_cell.value
                                if str(value1) == "None":
                                    value1 = ""
                                if str(value2) == "None":
                                    value2 = ""
                                if value1 == "" and value2 == "":
                                    blank_col_flag += 1
                                else:
                                    blank_col_flag = 0
                            else:
                                sheet1_cell.fill = PatternFill(start_color=self.Not_Agreed_color, end_color=self.Not_Agreed_color, fill_type="solid")
                                print(f"当前行数为：{inspect.currentframe().f_lineno}，value1 = {sheet1_cell.value}")
                                print(f"当前行数为：{inspect.currentframe().f_lineno}，value2 = {sheet2_cell.value}")

                        if blank_col_flag >= sheet1.max_column:
                            blank_row_flag += 1
                            if blank_row_flag >= 20:
                                print_info = f"连续20行全部值为空，结束sheet【{sheet_name}】的对比"
                                print(print_info)
                                comparisontask.signals.progress_current_task.emit(print_info)
                                break
                        else:
                            blank_row_flag = 0

            

        # 保存第一个工作簿，此时已包含对比和填充颜色后的结果
        try:
            print(f"saving file")
            comparisontask.signals.progress_current_task.emit(f"对比完成，文件保存中···")
            wb1.save(output_path)
            # comparisontask.signals.progress_current_task.emit(f"对比完成，File1保存成功")
            print(f"file saved")
            if (100 - self.Progress_percent) == (100%(len(wb1.sheetnames)+len(wb2.sheetnames))):
                self.Progress_percent = 100
                comparisontask.signals.progress_update.emit(self.Progress_percent)  #线程管理优化，采用信号槽方法
                self.Progress_percent = 0
                comparisontask.signals.progress_current_task.emit(f"对比完成，File2保存成功")
                return 1
            else:
                comparisontask.signals.progress_current_task.emit(f"对比完成，File1保存成功")
        except Exception as e:
            if isinstance(e, PermissionError):
                error = f"没有权限保存文件到指定路径，请检查文件权限设置。"
            elif isinstance(e, OSError) and "磁盘空间不足" in str(e):
                error = f"磁盘空间不足，无法保存文件，请清理磁盘空间后再试。"
            elif isinstance(e, FileNotFoundError):
                error = f"保存文件时文件路径不存在：{str(e)}"
                try:
                    os.mkdir(output_path_name.replace(".\\", ""))
                    error = f"文件夹 {output_path_name} 创建成功。"
                    wb1.save(output_path)
                    if (100 - self.Progress_percent) == (100%(len(wb1.sheetnames)+len(wb2.sheetnames))):
                        self.Progress_percent = 100
                        # ComparisonApp.compare_progress_bar.setValue(self.Progress_percent)
                        comparisontask.signals.progress_update.emit(self.Progress_percent)  #线程管理优化，采用信号槽方法
                        self.Progress_percent = 0
                    return 1
                except FileExistsError:
                    error = f"文件夹 {output_path_name} 已经存在。"
                except PermissionError:
                    error = f"没有权限创建文件夹 {output_path_name}。"
            else:
                error = f"保存文件时出现未知错误：{str(e)}"
            print(error)
            ctypes.windll.user32.MessageBoxW(None, error, "错误信息", 0x00000010)
            self.Progress_percent = 0
            comparisontask.signals.progress_current_task.emit(f"对比完成，File1保存成功")
            return 0
            
        return 1

    def check_thread_running(self, comparisontask):
        if not comparisontask.is_running:
            comparisontask.signals.comparison_finished.emit()
            comparisontask.signals.progress_current_task.emit(f"用户强制终止对比进程")
            return 1
        else:
            return 0

class Signals(QObject):
    progress_update = pyqtSignal(int)
    progress_current_task = pyqtSignal(str)
    comparison_finished = pyqtSignal()


# 自定义任务类，继承自 QRunnable
class ComparisonTask(QRunnable):
    def __init__(self, app):    #输入参数：ExcelComparisonApp类实例
        super().__init__()
        self.app = app
        self.is_running = True
        self.signals = Signals()

    def run(self):
        self.is_running = True
        if not self.app.perform_comparison(self):   #使用ExcelComparisonApp类实例下的perform_comparison方法
            self.app.set_button_status("Start")
            # return 0
        self.signals.progress_current_task.emit("/*************************************************结束任务*******************************************************/")
    def stop(self):
        self.is_running = False
        self.app.set_button_status("Start")
        return 0

class Stored_data:
    def __init__(self, number, index_column_number):
        if isinstance(number, int) and number > 0:
            self.file1_path = ""
            self.file2_path = ""
            self.sheet_name_edit = ["" for _ in range(number)]
            self.index_edit = [["" for _ in range(index_column_number)] for _ in range(number)]
            self.mapping_title_flag = [0 for _ in range(number)]
        else:
            print(f"number is not a int type or number > 0; number = {number}")

    def save_to_file(self, filename):
        # 将存储的数据保存到指定的 JSON 文件中
        # 构建一个包含所有要保存数据的字典
        # data = {
        #     "line_edit_text": self.line_edit_text,
        #     "check_box_state": self.check_box_state,
        #     "combo_box_index": self.combo_box_index
        # }
        # 以写入模式打开文件
        if not os.path.exists(os.path.dirname(filename)):   #os.path.dirname(filename)获取文件的路径，并检查路径是否存在，
            os.makedirs(os.path.dirname(filename))  #路径不存在，创建路径
        with open(filename, 'w') as f:
            # 使用 json.dump 方法将字典数据写入文件
            json.dump(self.__dict__, f)

    def load_from_file(self, filename):
        # 从指定的 JSON 文件中加载存储的数据
        if os.path.exists(filename) and os.path.getsize(filename) > 0:
            try:
                # 以读取模式打开文件
                with open(filename, 'r') as f:
                    # 使用 json.load 方法从文件中读取数据并转换为字典
                    data = json.load(f)
                    # 从字典中获取相应的数据并更新存储类的属性
                    self.__dict__.update(data)
                    print(f"self.current_data.mapping_title_flag = {self.mapping_title_flag}")
            except FileNotFoundError:
                # 如果文件不存在，忽略错误，保持默认值
                pass
        else:
            print("文件为空或不存在，跳过读取操作。")     

class ExcelComparisonApp(QWidget):
    def __init__(self):
        # 初始化图形界面应用。
        super().__init__()
        self.initUI()
        self.thread_pool = QThreadPool.globalInstance()
        self.CompareApp = Person_ComparisonApp(output_path_name)

    def initUI(self):
        #按钮颜色
        self.Button_Color = "#CECECE" #灰色
        self.Button_click_Color = "#079E61"
        self.Border_color = "#4480b2"
        # 设置图形界面布局和组件。
        self.layout = QVBoxLayout()
        self.layout.setAlignment(Qt.AlignTop)
        font = QFont()
        font.setBold(True)
        
        self.sheet_name_index_number = 5    # 可重点对比的sheet最大个数
        self.index_column_number = 2    # index组合对比的列数量，当前设置为可将2列组合作为索引
        self.start_flag = True
        
        self.current_data = Stored_data(self.sheet_name_index_number, self.index_column_number)   # 各部件当前状态数据
        self.sheet_name_edit = [0] * self.sheet_name_index_number
        self.major_sheet_layout = [0] * self.sheet_name_index_number
        self.index_edit = [[0 for _ in range(self.index_column_number)] for _ in range(self.sheet_name_index_number)]
        self.mapping_title_button = [0] * self.sheet_name_index_number
        self.mapping_title_button_func = [0] * self.sheet_name_index_number
        self.mapping_title_flag = [0] * self.sheet_name_index_number #flag代表：按索引对比的sheet的是否需要mapping表头，0:不需要，others：需要

        clear_button_layout = QHBoxLayout()
        self.clear_button = QPushButton("One-click clear")
        self.clear_button.setStyleSheet("""
            QPushButton {
                background-color: #CECECE;  /* 默认背景色为灰色 */
            }
            QPushButton:hover {
                background-color: #FFECA1;  /* 鼠标悬停时背景色变为淡蓝色 */
            }
        """f" border-style: solid; border-width: 1px; border-color: {self.Border_color}; border-radius:5")
        self.clear_button.clicked.connect(self.One_click_clear)
        self.clear_button.setFixedHeight(30)
        clear_button_layout.addWidget(self.clear_button)

        # 文件 1 选择部分
        file1_layout = QHBoxLayout()
        file1_label = QLabel("Select File 1：")
        self.file1_path_edit = QLineEdit()
        self.file1_path_edit.setStyleSheet(f""" border-style: solid; 
                                            border-width: 1px; 
                                            border-color: {self.Border_color}; 
                                            border-radius:3 """)
        file1_button = QPushButton("browse")
        file1_button.clicked.connect(self.select_file1)
        file1_layout.addWidget(file1_label)
        file1_layout.addWidget(self.file1_path_edit)
        file1_layout.addWidget(file1_button)

        # 文件 2 选择部分
        file2_layout = QHBoxLayout()
        file2_label = QLabel("Select File 2：")
        self.file2_path_edit = QLineEdit()
        self.file2_path_edit.setStyleSheet(f""" border-style: solid; 
                                            border-width: 1px; 
                                            border-color: {self.Border_color}; 
                                            border-radius:3 """)
        file2_button = QPushButton("browse")
        file2_button.clicked.connect(self.select_file2)
        file2_layout.addWidget(file2_label)
        file2_layout.addWidget(self.file2_path_edit)
        file2_layout.addWidget(file2_button)

        # 输入需要重点对比的sheet name及index列
        sheet_layout = QHBoxLayout()
        sheet_layout.setAlignment(Qt.AlignLeft)
        sheet_index_title_layout = QHBoxLayout()
        # check_button_layout = QHBoxLayout()
        sheet_label = QLabel("Sheets need to be compared by index")
        sheet_label.setFixedSize(800, 30)
        # sheet_label.setStyleSheet(f""" border-style: solid; 
        #                                 border-width: 1px; 
        #                                 border-color: {self.Border_color}; 
        #                                 border-radius:5 """)
        sheet_label.setFont(font)
        sheet_layout.addWidget(sheet_label)
        
        sheet_name_label = QLabel("sheet name")     #标题sheet名称
        sheet_index_label = QLabel("index column")  #标题 索引列
        sheet_index_blank_label = QLabel("Flag")    #Mapping Flag
        sheet_name_label.setAlignment(Qt.AlignCenter)
        sheet_index_label.setAlignment(Qt.AlignCenter)
        sheet_index_blank_label.setAlignment(Qt.AlignCenter)
        sheet_name_label.setFixedHeight(30)
        sheet_index_label.setFixedHeight(30)
        sheet_index_blank_label.setFixedHeight(30)
        sheet_name_label.setStyleSheet(f"""background-color: #87cbf0; border-radius:5 """)  #背景色+边框圆角
        sheet_index_label.setStyleSheet(f"""background-color: #87cbf0; border-radius:5 """) #背景色+边框圆角
        sheet_index_blank_label.setStyleSheet(f"""background-color: #87cbf0; border-radius:5 """) #背景色+边框圆角
        sheet_index_title_layout.addWidget(sheet_name_label, 5)
        sheet_index_title_layout.addWidget(sheet_index_label, 11)
        sheet_index_title_layout.addWidget(sheet_index_blank_label, 4)

        for index in range(0, self.sheet_name_index_number):
            self.major_sheet_layout[index] = QHBoxLayout()
            self.sheet_name_edit[index] = QLineEdit()
            self.sheet_name_edit[index].setStyleSheet(f""" border-style: solid; 
                                            border-width: 1px; 
                                            border-color: {self.Border_color}; 
                                            border-radius:2 """)
            for x in range(0, self.index_column_number):
                self.index_edit[index][x] = QLineEdit()
                self.index_edit[index][x].setStyleSheet(f""" border-style: solid; 
                                                border-width: 1px; 
                                                border-color: {self.Border_color}; 
                                                border-radius:2 """)
            #按钮：按照每列表头进行对比
            self.mapping_title_button[index] = QPushButton("Mapping Title")
            self.mapping_title_button[index].setStyleSheet("""
                QPushButton {
                    background-color: #FFFFFF;  /* 默认背景色为白色 */
                }
                QPushButton:hover {
                    background-color: #94c0e3;  /* 鼠标悬停时背景色变为淡蓝色 */
                }
            """f" border-style: solid; border-width: 1px; border-color: {self.Border_color}; border-radius:5")
            # self.mapping_title_button[index].setStyleSheet(f"background-color:{self.Button_Color};")
            self.mapping_title_button[index].clicked.connect(lambda _, idx=index: self.mapping_title_button_Func(idx))
            
            self.major_sheet_layout[index].addWidget(self.sheet_name_edit[index])
            for x in range(0, self.index_column_number):
                self.major_sheet_layout[index].addWidget(self.index_edit[index][x])
            self.major_sheet_layout[index].addWidget(self.mapping_title_button[index])
        
        # 当前正在执行的Task
        current_task_layout = QHBoxLayout()
        current_task_label = QLabel("Print Info")
        self.current_task_edit = QPlainTextEdit ()
        self.current_task_edit.setFixedHeight(70)
        # self.current_task_edit.setFixedWidth(600)
        self.current_task_edit.setStyleSheet(f""" border-style: solid; 
                                            border-width: 1px; 
                                            border-color: {self.Border_color}; 
                                            border-radius:3 """)
        self.current_task_edit.setReadOnly(True)
        current_task_layout.addWidget(current_task_label)
        current_task_layout.addWidget(self.current_task_edit)

        # 列出第一个文件的所有sheet name
        list_sheet_layout = QHBoxLayout()
        list_sheet_button = QPushButton("search all sheets in file 1")
        
        list_sheet_button.clicked.connect(self.list_sheet)
        list_sheet_layout.addWidget(list_sheet_button)
        self.list_sheet_name_layout = []
        self.list_sheet_name = []
        self.list_sheet_name_layout = []

        # 比较按钮
        compare_button_layout = QHBoxLayout()
        self.compare_button = QPushButton("Start Compare")
        self.compare_button.setStyleSheet("""
            QPushButton {
                background-color: #FFFFFF;  /* 默认背景色为白色 */
            }
            QPushButton:hover {
                background-color: #94c0e3;  /* 鼠标悬停时背景色变为淡蓝色 */
            }
        """f" border-style: solid; border-width: 1px; border-color: {self.Border_color}; border-radius:5")
        self.compare_button.setFixedHeight(40)
        # 比较进度显示
        self.compare_progress_bar = QProgressBar()
        self.compare_progress_bar.setMinimum(0)  # 设置进度条最小值为0
        self.compare_progress_bar.setMaximum(100)  # 设置进度条最大值为100
        # self.compare_button_status = QLineEdit()
        # self.compare_button_status.setFixedHeight(40)
        # self.compare_button_status.setEnabled(False) 
        # self.compare_progress_bar.setValue(i)
        self.compare_button.clicked.connect(self.start_comparison)
        compare_button_layout.addWidget(self.compare_button, 2)
        compare_button_layout.addWidget(self.compare_progress_bar, 3)

        # 空白部件
        blank_layout = QHBoxLayout()
        blank_widget = QWidget()
        blank_widget.setStyleSheet(f""" border-style: solid; 
                                        border-width: 1px; 
                                        border-color: {self.Border_color}; 
                                        border-radius:5 """)
        blank_layout.addWidget(blank_widget)
        # 空白部件
        blank2_layout = QHBoxLayout()
        blank2_widget = QWidget()
        blank2_widget.setStyleSheet(f""" border-style: solid; 
                                        border-width: 1px; 
                                        border-color: {self.Border_color}; 
                                        border-radius:5 """)
        blank2_layout.addWidget(blank2_widget)
        self.layout.addLayout(clear_button_layout)
        self.layout.addLayout(file1_layout)
        self.layout.addLayout(file2_layout)
        self.layout.addLayout(sheet_layout)
        # self.layout.addLayout(blank_layout)
        
        self.layout.addLayout(sheet_index_title_layout)
        for index in range(0, self.sheet_name_index_number):
            self.layout.addLayout(self.major_sheet_layout[index])
       
        self.layout.addLayout(compare_button_layout)
        # self.layout.addLayout(blank2_layout)
        self.layout.addLayout(current_task_layout)
        # 查找第一个文件的所有sheet name
        self.layout.addLayout(list_sheet_layout)
        # self.setLayout(self.layout)
        # self.setWindowTitle("Excel 文件对比工具"
        self.current_data.load_from_file('.\\json\\config.json')
        self.restore_current_data(self.current_data) #加载历史数据

    def One_click_clear(self):
        self.current_data.file1_path = ""
        self.current_data.file2_path = ""
        self.current_data.sheet_name_edit = [""] * self.sheet_name_index_number
        self.current_data.index_edit = [[""]*self.index_column_number] * self.sheet_name_index_number
        self.current_data.mapping_title_flag = [0] * self.sheet_name_index_number
        self.restore_current_data(self.current_data)

    def select_file1(self):
        # 选择第一个 Excel 文件，并在文本框中显示路径。
        file_path, _ = QFileDialog.getOpenFileName(self, "选择第一个 Excel 文件", "", "Excel 文件 (*.xlsx)")
        if file_path:
            self.file1_path_edit.setText(file_path)
            self.current_data.file1_path = file_path

    def select_file2(self):
        # 选择第二个 Excel 文件，并在文本框中显示路径。
        file_path, _ = QFileDialog.getOpenFileName(self, "选择第二个 Excel 文件", "", "Excel 文件 (*.xlsx)")
        if file_path:
            self.file2_path_edit.setText(file_path)
            self.current_data.file2_path = file_path
    
    def mapping_title_button_Func(self, index):  #----------------------------------------------------------------------------------------#
        # mapping表头的按钮状态切换
        if self.mapping_title_flag[index] != 0:
            self.mapping_title_flag[index] = 0
            self.current_data.mapping_title_flag[index] = 0
            self.mapping_title_button[index].setStyleSheet(f"background-color:{self.Button_Color};")
        else:
            self.mapping_title_flag[index] = 1
            self.current_data.mapping_title_flag[index] = 1
            self.mapping_title_button[index].setStyleSheet(f"background-color:{self.Button_click_Color};")
        print(f"self.mapping_title_flag = {self.mapping_title_flag}")

    def list_sheet(self):
        # 创建并显示第一个 Excel 文件文件所有的sheet名称的部件
        for index in range(0, len(self.list_sheet_name)):
            self.list_sheet_name_layout[index].removeWidget(self.list_sheet_name[index])
            print(f"删除部件{self.list_sheet_name[index].text()}")
        self.list_sheet_name = []
        self.list_sheet_name_layout = []
        print(f"list_sheet中，openning File 1")
        # self.current_task_edit.setText(f"list_sheet中，openning File 1")
        wb = self.CompareApp.open_file(self.file1_path_edit.text())
        if wb == 0:
            return 0
        
        self.sheet_name_layout = []*len(wb.sheetnames)
        for index in range(0, len(wb.sheetnames)):
            # self.current_task_edit.setText(f"list_sheet中，checking No.{index} group")
            print(f"list_sheet中，checking No.{index} group")
            self.list_sheet_name_layout.append(QHBoxLayout())
            self.list_sheet_name.append(QLineEdit())
            self.list_sheet_name[index].setText(wb.sheetnames[index])
            self.list_sheet_name_layout[index].addWidget(self.list_sheet_name[index])
            self.layout.addLayout(self.list_sheet_name_layout[index])
    
    # 检查index是否合法
    def check_name(self):
        # 清空compare_sheet_name
        self.compare_sheet_name = {}
        self.index_values = [[0 for _ in range(self.index_column_number+1)] for _ in range(self.sheet_name_index_number)]
        # 存放index的计算结果
        index_value = [0 for _ in range(self.sheet_name_index_number)]
        # 检查index是否合法
        for index in range(0, self.sheet_name_index_number):
            if self.sheet_name_edit[index].text() != '':
                all_null_flag = 0
                for x in range(0, self.index_column_number):
                    if self.index_edit[index][x].text() != '':
                        all_null_flag = 1
                if all_null_flag ==0:
                    print(f"all_null_flag ==0，跳过{self.sheet_name_edit[index].text()}sheet，不按索引对比")
                    continue

                name = self.sheet_name_edit[index].text()
                self.current_data.sheet_name_edit[index] = self.sheet_name_edit[index].text() #保存当前部件数据
                for x in range(0, self.index_column_number):    #index列增加至多个
                    index_value[x] = 0
                    try:
                        if len(self.index_edit[index][x].text()) >= 1 and self.index_edit[index][x].text().isalpha() :
                            print(f"INDEX全部都是字符串")
                            string = self.index_edit[index][x].text().upper() #字符串全部变为大写字母

                            for s in range(0, len(self.index_edit[index][x].text())):
                                index_arrays = ord(string[s]) - ord('A') + 1 #计算每个字母的值
                                index_value[x] = index_value[x] + index_arrays * (26 ** (len(string) - s - 1))
                                print(f"index_value = {index_value[x]}")
                        else:
                            if self.index_edit[index][x].text() != "":
                                index_value[x] = int(self.index_edit[index][x].text()) 
                                print(f"INDEX全部都是数字")
                            else:
                                print(f"第{x+1}组输入的索引值为空")
                        
                        self.current_data.index_edit[index][x] = self.index_edit[index][x].text() #保存当前部件数据
                        print(f"成功获取索引值：{index_value[x]}")
                        
                        self.index_values[index][x] = index_value[x]
                        if x >= self.index_column_number-1:
                            self.index_values[index][self.index_column_number] = self.mapping_title_flag[index]
                    except ValueError:
                        error = f"第{index+1}组输入的索引值不是有效的整数或字符，\n index = {self.index_values[index]}"
                        print(error)
                        ctypes.windll.user32.MessageBoxW(None, error, "错误信息", 0x00000010)
                        return 0
                self.compare_sheet_name[name] = self.index_values[index]
                print(f"self.compare_sheet_name[name][index] = {self.compare_sheet_name[name]}")
            else:
                self.current_data.sheet_name_edit[index] = ""
                for x in range(0, self.index_column_number):    #index列增加至多个
                    self.current_data.index_edit[index][x] = ""
                    if x >= self.index_column_number-1:
                        self.index_values[index][self.index_column_number] = self.mapping_title_flag[index]


                
        # print(f"compare_sheet_name{name} = {compare_sheet_name[name]}")
        return 1

    def start_comparison(self):
        # self.compare_button_status.setText("正在进行对比···")
        # 创建线程来执行文件打开和对比操作
        if self.start_flag:
            self.task = ComparisonTask(self)
            self.task.signals.progress_update.connect(self.compare_progress_bar.setValue)
            self.task.signals.progress_current_task.connect(self.current_task_edit.appendPlainText)
            self.task.signals.comparison_finished.connect(self.on_comparison_finished)
            self.thread_pool.start(self.task)
            # self.compare_button.clicked.connect(self.stop_comparison)
            self.set_button_status("Stop")
            self.compare_progress_bar.setValue(0)
            print("thread is Running, Button set to Stop Button.")
            # self.compare_button.setEnabled(False)
            # self.compare_button.clicked.connect(self.start_comparison)
        else:
            # self.compare_button_status.setText("正在进行对比···")
            # 创建线程来执行文件打开和对比操作
            self.task.stop()
            print(f"当前行数为：{inspect.currentframe().f_lineno}, self.task.stop()")
            # print(f"当前行数为：{inspect.currentframe().f_lineno}")
            self.thread_pool.clear()
            print(f"当前行数为：{inspect.currentframe().f_lineno}, self.thread_pool.clear()")
            self.thread_pool.waitForDone()
            print(f"当前行数为：{inspect.currentframe().f_lineno}, All tasks are stopped.")
            # self.compare_button.setEnabled(True)
            # self.set_button_status("Start")
            print(f"当前行数为：{inspect.currentframe().f_lineno}, thread is Stopped, Button set to Start Button.")

    def set_button_status(self, status):
        # self.compare_button_status.setText("正在进行对比···")
        # 创建线程来执行文件打开和对比操作
        if status == "Stop":
            self.start_flag = False
            self.compare_button.setEnabled(False)
            self.compare_button.setText("Stop")
            # print(f"当前行数为：{inspect.currentframe().f_lineno}")
            self.compare_button.setStyleSheet("""
                QPushButton {
                    background-color: #FFFFFF;  /* 默认背景色为红色 */
                }
                QPushButton:hover {
                    background-color: #D20103;  /* 鼠标悬停时背景色变为淡蓝色 */
                }
            """f" border-style: solid; border-width: 1px; border-color: {self.Border_color}; border-radius:5")

            self.compare_progress_bar.setStyleSheet("QProgressBar { color: Green; }")
            # print(f"当前行数为：{inspect.currentframe().f_lineno}")
            self.compare_button.setEnabled(True)
            print(f"当前行数为：{inspect.currentframe().f_lineno}, Button set to Stop Button.")
            # self.compare_button.setEnabled(False)
            # self.compare_button.clicked.connect(self.start_comparison)
        elif status == "Start":
            # self.compare_button_status.setText("正在进行对比···")
            # 创建线程来执行文件打开和对比操作
            self.start_flag = True
            self.compare_button.setEnabled(False)
            self.compare_button.setText("Start Compare")
            self.compare_button.setStyleSheet("""
                QPushButton {
                    background-color: #FFFFFF;  /* 默认背景色为白色 */
                }
                QPushButton:hover {
                    background-color: #94c0e3;  /* 鼠标悬停时背景色变为淡蓝色 */
                }
            """f" border-style: solid; border-width: 1px; border-color: {self.Border_color}; border-radius:5")
            # self.compare_button.clicked.connect(self.start_comparison)
            self.compare_button.setEnabled(True)
            # print(f"当前行数为：{inspect.currentframe().f_lineno}")
            # self.compare_progress_bar.setValue(0)
            print(f"当前行数为：{inspect.currentframe().f_lineno}, set to Start Button.")
        else:
            print(f"当前行数为：{inspect.currentframe().f_lineno}, set_button_status status input error, Start or Stop?")
            return 0
        return 1

    def on_comparison_finished(self):   #线程完成的槽函数，当进度达到100%时发射信号调用
        # self.compare_button.setEnabled(True)
        # self.thread_pool.waitForDone()
        self.set_button_status("Start")
        self.current_data.save_to_file('.\\json\\config.json')
        # self.stop_comparison()

    def perform_comparison(self, comparisontask):
        # 开始对比操作，获取输入文件和输出文件路径后调用对比函数。
        # global output_path_name
        # global wb1
        # global wb2
        # success_flag = True
        self.current_task_edit.clear()
        
        comparisontask.signals.progress_current_task.emit("/*********************************************************开始任务***********************************************/\nopenning File 1")
        print("openning WorkBook1")
        wb1 = self.CompareApp.open_file(self.file1_path_edit.text())
        if wb1 == 0:
            return 0
        print("openning WorkBook2")
        comparisontask.signals.progress_current_task.emit("openning File 2")
        wb2 = self.CompareApp.open_file(self.file2_path_edit.text())
        if wb2 == 0:
            return 0
        file1_name = (str(self.file1_path_edit.text().split('/')[-1])).rsplit('.', 1)
        file2_name = (str(self.file2_path_edit.text().split('/')[-1])).rsplit('.', 1)
        excel1_name = str(file1_name[0])
        excel2_name = str(file2_name[0])
        output_path1 = output_path_name + "\\" + excel1_name + "-compare.xlsx"
        output_path2 = output_path_name + "\\" + excel2_name + "-compare.xlsx"
                        
        if not self.check_name():
            # self.compare_button.setEnabled(True)
            # self.start_comparison()
            comparisontask.signals.progress_current_task.emit(f"index值校验失败，输入非法！")
            return 0
        # output_path_name = self.output_path_edit.text()
        comparisontask.signals.progress_current_task.emit("File1 Compare with File2")
        if not self.CompareApp.compare_excel_sheets_by_index(comparisontask, wb1, wb2, output_path1, self.compare_sheet_name, 0):
            # self.compare_button_status.setText("对比失败") 
            # self.compare_button.setEnabled(True)
            # self.set_button_status("Start")
            self.compare_progress_bar.setStyleSheet("QProgressBar { color: red; }")
            self.CompareApp.Progress_percent = 0
            # comparisontask.signals.progress_current_task.emit("File1 Compare with File2 Failed")
            return 0
        comparisontask.signals.progress_current_task.emit("File2 Compare with File1")
        if not self.CompareApp.compare_excel_sheets_by_index(comparisontask, wb2, wb1, output_path2, self.compare_sheet_name, 1):
            # self.compare_button_status.setText("对比失败") 
            # self.compare_button.setEnabled(True)
            # self.set_button_status("Start")
            self.compare_progress_bar.setStyleSheet("QProgressBar { color: red; }")
            self.CompareApp.Progress_percent = 0
            # comparisontask.signals.progress_current_task.emit("File2 Compare with File1 Failed")
            return 0
        # self.compare_button_status.setText("对比完成") 
        ctypes.windll.user32.MessageBoxW(None, "对比完成，输出文件在“outputfile”文件夹中", "成功信息", 0x00000040)
        comparisontask.signals.comparison_finished.emit()
        self.CompareApp.Progress_percent = 0
        return 1

    def restore_current_data(self, current_data):
        self.file1_path_edit.setText(current_data.file1_path)
        self.file2_path_edit.setText(current_data.file2_path)
        for index in range(0, self.sheet_name_index_number):
            self.sheet_name_edit[index].setText(current_data.sheet_name_edit[index])
            for x in range(0, self.index_column_number):
                self.index_edit[index][x].setText(current_data.index_edit[index][x])
            self.mapping_title_flag[index] = current_data.mapping_title_flag[index]
            if self.mapping_title_flag[index] != 0:
                self.mapping_title_button[index].setStyleSheet(f"background-color:{self.Button_click_Color};")
            else:
                self.mapping_title_button[index].setStyleSheet(f"background-color:{self.Button_Color};")

    def closeEvent(self, event):
        # 保存当前数据
        # self.current_data.save_to_file("config.json")
        # 等待线程池中的任务完成
        self.thread_pool.waitForDone()
        self.current_data.save_to_file('.\\json\\config.json')
        event.accept()
        
if __name__ == "__main__":
    # 主程序入口，创建应用实例并显示图形界面。
    app = QApplication(sys.argv)
    ex = ExcelComparisonApp()
    ex.setLayout(ex.layout)
    ex.setWindowTitle("Excel 文件对比工具")
    ex.show()
    sys.exit(app.exec_())

